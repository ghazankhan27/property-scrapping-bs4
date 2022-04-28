from calendar import c
import requests
from random_user_agent.user_agent import UserAgent
from random_user_agent.params import SoftwareName, OperatingSystem
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,GradientFill,PatternFill, Font


def prepareUrl(zip_code):
    wb = load_workbook(filename = 'uszips.xlsx')
    sheet = wb['Sheet1']

    zip = 0
    city=""
    state=""
    found = False


    firstrow = 2
    lastrow = 33789

    zip_col = 1
    city_col = 4
    state_col = 5

    for row in range(firstrow, lastrow+1):
        cell = sheet.cell(row=row, column=zip_col)
        if cell.value == zip_code:
            found = True
            zip = zip_code
            city = sheet.cell(row=row, column=city_col).value.lower().strip().replace(" ", "-")
            state = sheet.cell(row=row, column=state_col).value.lower()
            break


    
        
    if found:
        url_string = 'https://www.apartments.com/{0}-{1}-{2}/'.format(city,state,zip)

        print(url_string)
        wb.close()

        return url_string

    else:
        wb.close()
        return False


def add_to_xl(property_object, url):
    workbook = Workbook()
    sheet = workbook.active

    starting_col = 1

    temp1 = sheet.cell(column=starting_col, row=2, value="URL")
    temp1.fill = PatternFill("solid", fgColor="494949")
    temp1.font = Font(color='FFFFFF')
    sheet.cell(column=starting_col, row=3, value=url)

    temp1 = sheet.cell(column=starting_col+1, row=2, value="Name")
    temp1.fill = PatternFill("solid", fgColor="494949")
    temp1.font = Font(color='FFFFFF')
    sheet.cell(column=starting_col+1, row=3, value=property_object['name'])

    temp1 = sheet.cell(column=starting_col+2, row=2, value="Address")
    temp1.fill = PatternFill("solid", fgColor="494949")
    temp1.font = Font(color='FFFFFF')
    sheet.cell(column=starting_col+2, row=3, value=property_object['address'])

    temp1 = sheet.cell(column=starting_col+3, row=2, value="City")
    temp1.fill = PatternFill("solid", fgColor="494949")
    temp1.font = Font(color='FFFFFF')
    sheet.cell(column=starting_col+3, row=3, value=property_object['city'])

    temp1 = sheet.cell(column=starting_col+4, row=2, value="State")
    temp1.fill = PatternFill("solid", fgColor="494949")
    temp1.font = Font(color='FFFFFF')
    sheet.cell(column=starting_col+4, row=3, value=property_object['state'])

    temp1 = sheet.cell(column=starting_col+5, row=2, value="Built")
    temp1.fill = PatternFill("solid", fgColor="494949")
    temp1.font = Font(color='FFFFFF')
    sheet.cell(column=starting_col+5, row=3, value=property_object['built'])

    temp1 = sheet.cell(column=starting_col+6, row=2, value="Units")
    temp1.fill = PatternFill("solid", fgColor="494949")
    temp1.font = Font(color='FFFFFF')
    sheet.cell(column=starting_col+6, row=3, value=property_object['units'])
    

    len_avail_prop = len(property_object['available_properties'])
    len_unavail_prop = len(property_object['unavailable_properties'])
    single_prop = property_object['property']
    len_avail_range_prop = len(property_object['available_properties_range'])

    col_counter = starting_col+7

    if len_avail_prop > 0:
        j = col_counter

        for col in range(j, j + len_avail_prop):
            key = list(property_object['available_properties'][col-j].keys())[0]

            prop_array = property_object['available_properties'][col-j][key]

            sheet.merge_cells(start_row=1, start_column=col_counter , end_row=1, end_column=col_counter+2)

            x = sheet.cell(row=1, column=col_counter, value=key)
            x.alignment = Alignment(horizontal='center')
            x.fill = PatternFill("solid", fgColor="494949")
            x.font = Font(color='FFFFFF')  
            
            temp1 = sheet.cell(column=col_counter, row=2, value="Price")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')

            temp1 = sheet.cell(column=col_counter+1, row=2, value="SF")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')
            
            temp1 = sheet.cell(column=col_counter+2, row=2, value="Price/SF")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')

            sum = 0

            for i in range(3,len(prop_array)+3):
                price = str(prop_array[i-3]['price']).strip()
                try:
                    price_int = int(price.replace("$","").replace(",","").replace("*",""))
                    sum += price_int
                except:
                    continue
                units = int(str(prop_array[i-3]['area']).replace(",",""))
                avg = "$" + str(round(price_int/units,2))


                sheet.cell(column=col_counter, row=i, value=price)
                sheet.cell(column=col_counter+1, row=i, value=units)
                sheet.cell(column=col_counter+2, row=i, value=avg)

            avg_price = sum/len(prop_array)
            x = sheet.cell(column=col_counter, row=len(prop_array)+5, value="$" + str(avg_price))
            x.fill = PatternFill("solid", fgColor="0094E0")
            x.font = Font(color='FFFFFF')  
            
            col_counter+=3

    if len_unavail_prop > 0:
        
        j = col_counter

        prop_array = property_object['unavailable_properties']

        for col in range(j, j + len_avail_prop):

            key = prop_array[col-j]

            sheet.merge_cells(start_row=1, start_column=col_counter , end_row=1, end_column=col_counter+2)

            x = sheet.cell(row=1, column=col_counter, value=key)
            x.alignment = Alignment(horizontal='center')  
            x.fill = PatternFill("solid", fgColor="494949")
            x.font = Font(color='FFFFFF')
            
            temp1 = sheet.cell(column=col_counter, row=2, value="Price")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')

            temp1 = sheet.cell(column=col_counter+1, row=2, value="SF")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')
            
            temp1 = sheet.cell(column=col_counter+2, row=2, value="Price/SF")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')


            col_counter+=3

    if single_prop:
        key = list(property_object['property'].keys())[0]

        sheet.merge_cells(start_row=1, start_column=col_counter , end_row=1, end_column=col_counter+2)

        x = sheet.cell(row=1, column=col_counter, value=key)
        x.alignment = Alignment(horizontal='center')
        x.fill = PatternFill("solid", fgColor="494949")
        x.font = Font(color='FFFFFF')  

        temp1 = sheet.cell(column=col_counter, row=2, value="Price")
        temp1.fill = PatternFill("solid", fgColor="494949")
        temp1.font = Font(color='FFFFFF')

        temp1 = sheet.cell(column=col_counter+1, row=2, value="SF")
        temp1.fill = PatternFill("solid", fgColor="494949")
        temp1.font = Font(color='FFFFFF')

        temp1 = sheet.cell(column=col_counter+2, row=2, value="Price/SF")
        temp1.fill = PatternFill("solid", fgColor="494949")
        temp1.font = Font(color='FFFFFF')


        price = str(single_prop[key]).strip()
        price_int = int(price.replace("$","").replace(",","").replace("*",""))
        units = int(str(single_prop['units']).replace(",",""))
        
        avg = "$" + str(round(price_int/units,2))


        sheet.cell(column=col_counter, row=3, value=price)
        sheet.cell(column=col_counter+1, row=3, value=units)
        sheet.cell(column=col_counter+2, row=3, value=avg)

        col_counter+=3
        
    if len_avail_range_prop > 0:

        props = property_object['available_properties_range']

        sum = 0

        for prop in props:


            key = list(prop.keys())[0]

            sheet.merge_cells(start_row=1, start_column=col_counter , end_row=1, end_column=col_counter+2)

            temp1 = sheet.cell(column=col_counter, row=1, value=key)
            temp1.alignment = Alignment(horizontal="center")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')

            temp1 = sheet.cell(column=col_counter, row=2, value="Price")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')

            temp1 = sheet.cell(column=col_counter+1, row=2, value="SF")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')

            temp1 = sheet.cell(column=col_counter+2, row=2, value="Price/SF")
            temp1.fill = PatternFill("solid", fgColor="494949")
            temp1.font = Font(color='FFFFFF')

            price = int(str(prop[key]).replace("$", "").replace("*","").replace(",","").strip())
            sum += price
            units = int(str(prop['units']).replace(",","").strip())
            avg = "$" + str(round(price/units,2))

            sheet.cell(column=col_counter, row=3, value=str(price)+"*")
            sheet.cell(column=col_counter+1, row=3, value=units)
            sheet.cell(column=col_counter+2, row=3, value=avg)


            col_counter+=3


    file_name_string = "property-{0}.xlsx".format(property_object['name'])

    workbook.save(filename=file_name_string)
    workbook.close()


def get_random_user_agent():
    software_names = [SoftwareName.CHROME.value]
    operating_systems = [OperatingSystem.WINDOWS.value, OperatingSystem.LINUX.value]   

    user_agent_rotator = UserAgent(software_names=software_names, operating_systems=operating_systems, limit=100)

    # Get list of user agents.
    user_agents = user_agent_rotator.get_user_agents()

    # Get Random User Agent String.
    user_agent = user_agent_rotator.get_random_user_agent()

    return user_agent


def get_all_urls(soup):

    properties_list = soup.find_all('li', class_='mortar-wrapper')

    urls_list = []


    for property in properties_list:
        url = property.find('article')['data-url']
        urls_list.append(url)
    

    return urls_list


def get_all_properties_range_price(properties_list):
    empty_properties_list = []

    length = int(int(len(properties_list)) / 2)

    for i in range(length):
        property = properties_list[i]
        temp = property.find('div', class_='row').find('div','column1').find('h4','detailsLabel').find('span', class_='detailsTextWrapper').find_all('span')
        temp_price = property.find('div', class_='row').find('div','column1').find('h3','modelLabel').find('span', class_='rentLabel')
        temp1 = str(temp[0].text)
        temp2 = str(temp[1].text)
        temp3 = str(temp[2].text).replace("units","").replace("sq ft", "").strip()
        temp_price = str(temp_price.text).strip()
        temp_price = temp_price.split()
        temp_price = temp_price[0] + "*"
        empty_properties_list.append({
            str(temp1+ " " +temp2): temp_price,
            "units":temp3
            })

    return empty_properties_list


def get_all_properties(properties_list):

    empty_properties_list = []

    length = int(int(len(properties_list)) / 2)

    for i in range(length):
        property = properties_list[i]
        temp = property.find('div', class_='row').find('div','column1').find('h4','detailsLabel').find('span', class_='detailsTextWrapper').find_all('span')
        temp1 = str(temp[0].text)
        temp2 = str(temp[1].text)
        empty_properties_list.append(str(temp1+ " " +temp2))

    return empty_properties_list


def get_all_properties_grid(properties_list):

    all_properties_details = []


    length = int(int(len(properties_list)) / 2)

    for i in range(length):
        price_objects = []
        property = properties_list[i]
        temp = property.find('div', class_='row').find('div','column1').find('h4','detailsLabel').find('span', class_='detailsTextWrapper').find_all('span')
        units_grid = property.select_one('div.unitGridContainer.mortar-wrapper').find('ul').find_all('li')


        for unit in units_grid:           
            temp_price_obj = {}
            price = unit.select_one('div.pricingColumn.column').find_all('span')
            sq_foot = unit.select_one('div.sqftColumn.column').find_all('span')
            price = str(price[1].text)
            sq_foot = str(sq_foot[1].text)

            
            temp_price_obj['price'] = price.strip()
            temp_price_obj['area'] = sq_foot.strip()

            price_objects.append(temp_price_obj)

        temp1 = str(temp[0].text)
        temp2 = str(temp[1].text)

        all_properties_details.append({
        str(temp1+ " " +temp2) : price_objects
        })
    
    return all_properties_details
    

def get_single_property(property):

    temp_price = property.find('div', class_='row').find('div','column1').find('h3','modelLabel').find('span', class_='rentLabel').text
    temp_description = property.find('div', class_='row').find('div','column1').find('h4','detailsLabel').find('span', class_='detailsTextWrapper').find_all('span')
    temp_price = str(temp_price).strip()
    temp1 = str(temp_description[0].text)
    temp2 = str(temp_description[1].text)
    temp3 = str(temp_description[2].text).replace("units","").replace("sq ft","").strip()

    description_string = temp1 + " " + temp2

    return {
        description_string:temp_price,
        "units":temp3
    }


def get_all_data(url):
    temp_obj = {}
    headers = {'user-agent': get_random_user_agent()} # Setting a random user agent

    try:
        page_source = requests.get(url, headers=headers).text # Getting page source
        soup = BeautifulSoup(page_source, 'lxml') # Creating the soup
    except:
        print("Error in getting the page source")
        return


    try:
        property_name = str(soup.find('h1', id='propertyName').text).strip() # Getting the property name
        temp_obj['name'] = property_name    
    except:
        print("Error getting the property name")
        return

    try:
        # Getting the address string

        property_address_items = soup.find(id='propertyAddressRow').find(class_='propertyAddressContainer').find('h2').find_all('span')
        state = str(property_address_items[3].text)
        zip = str(property_address_items[4].text)
        property_address = str(property_address_items[0].text) + ", " + str(property_address_items[1].text) + ", " + state + ", " + zip
        temp_obj['address'] = property_address

    except:
        print("Error getting the address")
        return


    try:
        city = str(property_address_items[5].find('a').text) # Getting the city name
        temp_obj['city'] = city

    except:
        print("Error getting the city name")
        return


    try:
        state_string = soup.find('div', id='breadcrumbs-container').select('span.crumb')
        state = str(state_string[1].find('a').text).strip()
        temp_obj['state'] = state

    except:
        print("No state found")
        temp_obj['state'] = ""

    try:
        built_units_info = soup.select('div.mortar-wrapper.feesPoliciesCard.with-bullets-card')
        built_units_info = built_units_info[-1]
        built_units_info = built_units_info.find_all('li', class_='with-bullets')
        built_info = built_units_info[0].find('div', class_='column')
        unit_count = built_units_info[1].find('div', class_='column')
        built_info = str(built_info.text).replace('Built in ', "")
        unit_count = str(unit_count.text).replace("units","").replace("sq ft","").strip()
        temp_obj['built'] = built_info
        temp_obj['units'] = unit_count



    except:
        print("Error getting built and unit info")
        temp_obj['built'] = ""
        temp_obj['units'] = ""


    try:
        pricing_info = soup.find('div', id='pricingView') # Get the pricing container

        temp_obj['available_properties'] = []
        temp_obj['unavailable_properties'] = []
        temp_obj['property'] = []
        temp_obj['available_properties_range'] = []

        all_properties_grid = pricing_info.select('div.pricingGridItem.multiFamily.hasUnitGrid') # Get all available properties
        if len(all_properties_grid) > 0:
            available_properties = get_all_properties_grid(all_properties_grid)
            temp_obj['available_properties'] = available_properties
        else:
            temp_obj['available_properties'] = []

        

        if temp_obj['available_properties'] != []:
            all_properties = pricing_info.select('div.pricingGridItem.multiFamily') # Get all unavailable properties
            if len(all_properties) > 0:
                unavailable_properties = get_all_properties(all_properties)
                temp_obj['unavailable_properties'] = unavailable_properties
            else:
                temp_obj['unavailable_properties'] = []
        

        
        else:
            all_properties = pricing_info.select('div.pricingGridItem.multiFamily') # Get all unavailable properties
            if len(all_properties) > 0:
                available_properties = get_all_properties_range_price(all_properties)
                temp_obj['available_properties_range'] = available_properties
            else:
                temp_obj['available_properties_range'] = []

        

        
        try:
            if (len(temp_obj['available_properties']) <=0 and len(temp_obj['unavailable_properties']) <= 0 and len(temp_obj['available_properties_range']) <= 0):
                single_property = pricing_info.find('div', class_='pricingGridItem')
                property_details = get_single_property(single_property)
                temp_obj['property'] = property_details
            
        except:
            temp_obj['property'] = {}

             

    except:
        print("Error getting price objects")
        return

    return temp_obj


def scrape_site(url):

    headers = {'user-agent': get_random_user_agent()} # Setting a random user agent

    try:
        page_source = requests.get(url, headers=headers).text # Getting page source


        soup = BeautifulSoup(page_source, 'lxml') # Creating the soup

        urls_list = get_all_urls(soup)

        

        for i in range(len(urls_list)):
            print(i,urls_list[i])
        

        while True:
            selected_url = int(input("Make a selection (Enter -1 to end program): "))

            if selected_url == -1:
                break

            try:
                data_obj = get_all_data(urls_list[selected_url])
            except:
                print("Unable to get data")


            try:
                add_to_xl(data_obj,urls_list[selected_url])
            except:
                print("Error adding to excel sheet")


    except:
        print("Error getting page source")
    

def main():
    zip = int(input("Enter the zip code of your choice: "))

    url = prepareUrl(zip)

    if url != False:
        scrape_site(url)
    else:
        print("Invalid Zip Code")
    

main()